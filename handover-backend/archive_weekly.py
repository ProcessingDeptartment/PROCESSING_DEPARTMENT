#!/usr/bin/env python3
"""
Weekly handover archive script.
Downloads the weekly export from the backend and appends it to a master Excel file.

Usage:
    python archive_weekly.py

Configuration (edit these):
    BACKEND_URL = URL of the handover backend service
    ARCHIVE_PATH = Local path to save the master Excel file
"""

import requests
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# Configuration
BACKEND_URL = "https://handover-backend.onrender.com"
ARCHIVE_PATH = r"C:\Users\michaelac\OneDrive - Abagold\6. Daily meetings & Handovers\End of shift report\productionhandover.xlsx"

def get_week_range():
    """Returns (monday_str, sunday_str) for the previous complete week."""
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday.strftime('%Y-%m-%d'), sunday.strftime('%Y-%m-%d')

def download_export():
    """Download the weekly export from backend."""
    url = f"{BACKEND_URL}/api/handovers/export"
    print(f"Downloading from {url}...")
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return resp.content

def append_to_master(xlsx_bytes):
    """Append downloaded data to the master Excel file."""
    monday, sunday = get_week_range()
    sheet_name = f"week_{monday}_to_{sunday}"

    # Load downloaded file into memory
    from io import BytesIO
    from openpyxl import load_workbook as load_xlsx

    downloaded_wb = load_xlsx(BytesIO(xlsx_bytes))
    downloaded_ws = downloaded_wb.active

    # Load or create master file
    archive_dir = os.path.dirname(ARCHIVE_PATH)
    os.makedirs(archive_dir, exist_ok=True)

    if os.path.exists(ARCHIVE_PATH):
        master_wb = load_workbook(ARCHIVE_PATH)
        # Remove sheet if it already exists this week
        if sheet_name in master_wb.sheetnames:
            del master_wb[sheet_name]
    else:
        master_wb = Workbook()
        master_wb.remove(master_wb.active)  # Remove default sheet

    # Copy downloaded sheet to master
    new_ws = master_wb.create_sheet(sheet_name, 0)  # Insert at beginning
    for row in downloaded_ws.iter_rows(values_only=True):
        new_ws.append(row)

    # Format header row
    header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for cell in new_ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font

    # Auto-fit columns
    for col in new_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        new_ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Save master file
    master_wb.save(ARCHIVE_PATH)
    print(f"✓ Archived week {monday} to {sunday}")
    print(f"✓ Saved to {ARCHIVE_PATH}")

if __name__ == '__main__':
    try:
        print("Handover Weekly Archive")
        print("-" * 40)
        xlsx_bytes = download_export()
        append_to_master(xlsx_bytes)
        print("✓ Done!")
    except Exception as e:
        print(f"✗ Error: {e}")
        exit(1)
