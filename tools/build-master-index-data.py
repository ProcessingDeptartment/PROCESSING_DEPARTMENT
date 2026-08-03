"""
Regenerates public/lib/master-index-data.js from the controlled Master Index List
spreadsheet. Run this -- do not hand-edit the generated file -- whenever a new
revision of the index is issued, then update XL below to the new filename.

    python tools/build-master-index-data.py

Requires openpyxl.
"""
import openpyxl, re, os, glob, json, datetime

XL = r'T:\Abagold Processing Facility\14. Projects\20. Paperless\6. RECORDS\FINAL\_____REC 01 Master Index List_06.2026.xlsx'
RECDIR = r'T:\Abagold Processing Facility\14. Projects\20. Paperless\PROCESSING_DEPARTMENT\public\records'
OUT = os.path.join(RECDIR, '..', 'lib', 'master-index-data.js')

def norm(c):
    return re.sub(r'[^A-Z0-9]', '', str(c).upper())

def clean(v):
    if v is None: return ''
    if isinstance(v, datetime.datetime): return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    s = s.replace('\ufffd', '-').replace('\u2013', '-').replace('\u2014', '-')
    s = re.sub(r'\s+', ' ', s)
    return s

def cleandate(v):
    if v is None: return ''
    if isinstance(v, datetime.datetime): return v.strftime('%d/%m/%Y')
    s = clean(v).replace('.', '/')
    return s

def cleanyn(v):
    s = clean(v).upper()
    if s in ('Y', 'YES'): return 'Y'
    if s in ('N', 'NO'): return 'N'
    if s.startswith('N/A'): return 'N/A'
    return s

# --- built pages: docCode -> (recordKey, href)
pages = {}
for p in sorted(glob.glob(os.path.join(RECDIR, '*.html'))):
    fn = os.path.basename(p)
    if fn in ('master-record-index.html', 'record-list.html', 'batch-trace.html'): continue
    t = open(p, encoding='utf-8').read()
    m = re.search(r"docCode:\s*'([^']*)'", t)
    rk = re.search(r"recordKey:\s*'([^']*)'", t)
    if m:
        key = rk.group(1) if rk else fn[:-5]
    else:
        m = re.search(r'<span class="doc-code">([^<]*)</span>', t)
        if not m: continue
        key = fn[:-5]
    pages.setdefault(norm(m.group(1)), []).append((key, fn))

# Master-index document numbers that are written differently in the built pages.
ALIAS = {
    'REC7741C': 'REC774C',   # index "REC 7.7.4.1.c" vs page "REC 7.7.4c" (Knives Register)
    'QA01': 'QA1',           # index "QA 01" vs page "QA1" (Damaged Cans and Lids)
}

wb = openpyxl.load_workbook(XL, data_only=True)
ws = wb['REC']

# header block (rows 1-5)
hdr = {}
for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    cells = [clean(c) for c in r]
    for i, c in enumerate(cells):
        if c.endswith(':') or c in ('Page',):
            hdr[c.rstrip(':')] = cells[i+1] if i+1 < len(cells) else ''

rows = []
unlinked = []
# Pages already spoken for. Two index rows can carry the same document number -- REC 7.2.10
# is both "Stock Loading Control Sheet" and a superseded "Brine muixing report" (the live one
# is REC 7.3.6). Matching is by document number, so without this the second row inherits the
# first row's page and the index quietly links one record to a different controlled document.
claimed = set()
for r in ws.iter_rows(min_row=7, values_only=True):
    code, name, details, rev, obs, dist, issued = (list(r) + [None]*7)[:7]
    code, name = clean(code), clean(name)
    if not code and not name: continue
    if code.lower() == 'records' and not name: continue   # section separator
    if code == 'Document No.': continue                    # repeated column-header row
    try: revn = int(str(rev).strip())
    except: revn = None
    k = norm(code)
    k = ALIAS.get(k, k)
    hit = pages.get(k)
    # disambiguate the one duplicate docCode (REC 7.8.1 canning vs dry chiller)
    if hit and len(hit) > 1:
        hit = [h for h in hit if 'dry' not in h[0]] or hit
    # An unlinked row is honest ("not digitised"); a wrongly linked one is not.
    if hit and hit[0][0] in claimed:
        hit = None
    if hit:
        claimed.add(hit[0][0])
    rows.append({
        'docNo': code,
        'name': name,
        'details': clean(details),
        'revision': revn,
        'obsoleteRetrieved': cleanyn(obs),
        'distributed': cleanyn(dist),
        'dateOfIssue': cleandate(issued),
        'recordKey': hit[0][0] if hit else None,
        'href': hit[0][1] if hit else None,
    })
    if not hit and revn is not None:
        unlinked.append(code)

# The spreadsheet splits one record across two rows: the name sits on a row with no
# document number, and the number ("QA 01") on the next row with no name. Merge them
# so the record has both and can be matched to its built page.
merged = []
for r in rows:
    if merged and not r['name'] and r['docNo'] and merged[-1]['name'] and not merged[-1]['docNo']:
        prev = merged.pop()
        r = dict(r, name=prev['name'])
        k = norm(r['docNo'])
        hit = pages.get(ALIAS.get(k, k))
        if hit:
            r['recordKey'], r['href'] = hit[0][0], hit[0][1]
    merged.append(r)
rows = merged

# built pages that the index has no row for -> appended as facility-local records
listed = {r['recordKey'] for r in rows if r['recordKey']}
extra = []
for k, v in pages.items():
    for key, fn in v:
        if key not in listed:
            extra.append((key, fn))

# Built pages the controlled index has no row for. They must still appear, otherwise
# the online Master Record Index silently omits 21 live records.
for key, fn in sorted(extra):
    t = open(os.path.join(RECDIR, fn), encoding='utf-8').read()
    mc = re.search(r"docCode:\s*'([^']*)'", t) or re.search(r'<span class="doc-code">([^<]*)</span>', t)
    mt = re.search(r"\n\s*title:\s*'([^']*)'", t) or re.search(r'<title>([^<]*)</title>', t)
    mr = re.search(r'docRevisionStart:\s*(\d+)', t) or re.search(r'<span class="doc-rev">Rev (\d+)', t)
    name = clean(mt.group(1)) if mt else key
    code = clean(mc.group(1)) if mc else ''
    # Strip a leading document number from the page title, but only a real one --
    # 'Master Cleaning Checklist (New Plant)' must not lose its whole name.
    if code and re.match(r'^(REC|QA)', code) and name.startswith(code):
        name = name[len(code):].strip()
    rows.append({
        'docNo': code, 'name': name,
        'details': '', 'revision': int(mr.group(1)) if mr else 1,
        'obsoleteRetrieved': '', 'distributed': '', 'dateOfIssue': '',
        'recordKey': key, 'href': fn, 'notOnIndexList': True,
    })

js = """/*
 * Baseline data for REC 01 Master Index List, transcribed from
 * "6. RECORDS/FINAL/_____REC 01 Master Index List_06.2026.xlsx" (sheet REC,
 * Rev %s, revision date %s).
 *
 * This is the PAPER baseline only. The live current revision of any row that has
 * a `recordKey` comes from document-revision.js at render time -- never read
 * `revision` here as the current value for a digitised record.
 *
 * Regenerate rather than hand-edit when a newer Master Index List is issued.
 */
window.MasterIndexData = {
  header: %s,
  rows: %s
};
""" % (
    hdr.get('Revision', ''),
    hdr.get('Revision Date', ''),
    json.dumps({
        'document': hdr.get('Document', 'Master Index List'),
        'docNumber': hdr.get('Doc number', 'REC 01'),
        'reviewedBy': hdr.get('Reviewed by', ''),
        'approvedBy': hdr.get('Approved by', ''),
        'revision': hdr.get('Revision', ''),
        'revisionDate': hdr.get('Revision Date', ''),
        'effectiveDate': hdr.get('Effective Date', ''),
        'source': '_____REC 01 Master Index List_06.2026.xlsx',
    }, indent=4),
    json.dumps(rows, indent=4)
)
with open(OUT, 'w', encoding='utf-8', newline='') as f:
    f.write(js)

print('header:', hdr)
print('rows written:', len(rows))
print('linked to a page:', sum(1 for r in rows if r['recordKey']))
print('no page yet:', sum(1 for r in rows if not r['recordKey']))
print('\nbuilt pages NOT in the index (%d):' % len(extra))
for k, fn in sorted(extra): print('  ', k)
