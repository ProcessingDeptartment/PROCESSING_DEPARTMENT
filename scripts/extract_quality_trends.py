#!/usr/bin/env python3
"""
Extract the Quality department trending workbooks into a single JSON file the
FSMS Quality Trends dashboard reads (public/data/quality-trends.json).

Source workbooks live on the shared drive:
  Online system/19. Quality report/*.xlsx

Run after the quality team updates any of those workbooks:
  python scripts/extract_quality_trends.py

Read-only: nothing here writes back to the workbooks.
"""

import json
import os
import re
import sys
from datetime import datetime, date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.normpath(os.path.join(ROOT, '..', 'Online system', '19. Quality report'))
OUT = os.path.join(ROOT, 'public', 'data', 'quality-trends.json')

F_TEMP = 'Temperature, pH & Humidity Monitoring Trending 2025.xlsx'
F_ALLERGEN = 'Allergen Monitoring 2025.xlsx'
F_ENV = 'Environmental Monitoring Results 2025.xlsx'
F_QP = 'Quality Parameter Trending_Monthly Report 2025.xlsx'
F_EMP = 'REC 2.5.7.1 EMP Sampling Schedule Log and Trending.xlsx'

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']

# The workbooks are typed by hand, so both dates and readings contain slips
# (2042 pack dates, a vacuum of 1.5e21). Anything outside these windows is
# dropped from the dashboard and counted in data['excluded'] instead of
# silently skewing a trend line.
MIN_DATE = date(2015, 1, 1)
MAX_DATE = date.today()          # nothing can be measured in the future

PLAUSIBLE = {
    'ph': (2.0, 12.0),
    'salt': (0.0, 15.0),
    'brix': (0.0, 60.0),
    'vacuum': (0.0, 120.0),
    'drain': (0.0, 2000.0),
    'nett': (0.0, 5000.0),
}

EXCLUDED = {'dates': 0, 'readings': 0}


# ---------------------------------------------------------------- helpers

def norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def num(v):
    """
    A reading. QC often records several cans in one cell ("232/232",
    "210, 214", "20/30KPA") -- those are averaged rather than mangled into one
    huge number.
    """
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    if v is None:
        return None
    parts = [p for p in re.split(r'[\/,;&+]|\s+and\s+', str(v)) if p.strip()]
    vals = []
    for p in parts:
        # broken formula fragments left in cells ("5.9+CC28:C49") are not readings
        if re.match(r'^\s*\$?[A-Za-z]{1,3}\$?\d+(\s*:.*)?\s*$', p):
            continue
        s = re.sub(r'[^0-9.\-]', '', p)
        if s in ('', '-', '.', '-.'):
            continue
        try:
            vals.append(float(s))
        except ValueError:
            continue
    if not vals:
        return None
    return round(sum(vals) / len(vals), 4)


def parse_date(v):
    """Dates here are hand typed: 05.01.2026, 5/1/2026, 17.1.2025 ..."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm(v)
    m = re.match(r'^(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})$', s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    if mo > 12 and d <= 12:      # transposed entry
        d, mo = mo, d
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return None
    try:
        out = date(y, mo, d)
    except ValueError:
        return None
    if not (MIN_DATE <= out <= MAX_DATE):
        EXCLUDED['dates'] += 1
        return None
    return out


def iso(d):
    return d.isoformat() if d else None


def sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def load(fname):
    path = os.path.join(SRC, fname)
    if not os.path.exists(path):
        print('  ! missing: %s' % fname, file=sys.stderr)
        return None, None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    stamp = datetime.fromtimestamp(os.path.getmtime(path)).isoformat(timespec='seconds')
    return wb, stamp


def stats(points, keys):
    out = {}
    for k in keys:
        vals = [p['v'][k] for p in points if p['v'].get(k) is not None]
        if not vals:
            continue
        out[k] = {
            'n': len(vals),
            'min': round(min(vals), 3),
            'max': round(max(vals), 3),
            'avg': round(sum(vals) / len(vals), 3),
            'last': vals[-1],
        }
    return out


# ------------------------------------------------- block-layout monitoring

def parse_blocks(ws, value_labels):
    """
    These monitoring sheets repeat a month block across the columns:
        Month | Date | <value cols...> | (blank spacer)
    Returns a flat, date-sorted list of {d, v:{key:number}}.
    """
    rows = sheet_rows(ws)
    header_i = None
    for i, row in enumerate(rows[:6]):
        if any(norm(c).lower() == 'month' for c in row):
            header_i = i
            break
    if header_i is None:
        return []
    header = rows[header_i]
    starts = [c for c, v in enumerate(header) if norm(v).lower() == 'month']

    points = {}
    for c0 in starts:
        cols = []
        for c in range(c0 + 2, len(header)):
            label = norm(header[c])
            if not label or label.lower() in ('month', 'date'):
                break
            cols.append((c, label))
        cols = [(c, l) for c, l in cols if l in value_labels]
        if not cols:
            continue
        for row in rows[header_i + 1:]:
            if c0 + 1 >= len(row):
                continue
            d = parse_date(row[c0 + 1])
            if not d:
                continue
            vals = {}
            for c, label in cols:
                n = num(row[c]) if c < len(row) else None
                if n is not None:
                    vals[value_labels[label]] = n
            if vals:
                points.setdefault(d, {}).update(vals)
    return [{'d': iso(d), 'v': points[d]} for d in sorted(points)]


def monitoring_trend(ws, tid, title, unit, value_labels, series, note=''):
    pts = parse_blocks(ws, value_labels)
    return {
        'id': tid,
        'title': title,
        'category': 'Monitoring',
        'chart': 'line',
        'unit': unit,
        'series': series,
        'points': pts,
        'stats': stats(pts, [s['key'] for s in series]),
        'note': note,
    }


# ------------------------------------------------------- quality parameters

QP_MAP = {
    'JOB NO': 'job', 'AG CODES': 'ag', 'PRODUCT DESCRIPTION': 'product',
    'PIECES': 'pieces', 'INGO WEIGHT (G)': 'ingot', 'NETT MASS (G)': 'nett',
    'DRAIN MASS (G)': 'drain', 'DRAIN MASS': 'drain', 'PH': 'ph',
    'SALT %': 'salt', 'SALT%': 'salt', 'BRIX': 'brix', 'VACUUM': 'vacuum',
    'QC ISSUES': 'issues', 'ISSUES': 'issues',
}
QP_DEFAULT_ORDER = ['job', 'ag', None, None, 'product', 'pieces', 'ingot',
                    'nett', 'drain', 'ph', 'salt', 'brix', 'vacuum', 'issues']


def ag_date(code, smo, syr):
    """
    AG codes carry the pack date: AG040826 -> 04.08.2026. The month/year half is
    mistyped often enough (AG260606 on the June 2026 sheet) that the sheet the
    row lives on wins, and only the day is taken from the code.
    """
    m = re.match(r'^AG(\d{6})$', norm(code).upper().replace(' ', ''))
    if not m or not smo:
        return None
    day = int(m.group(1)[0:2])
    try:
        return date(syr, smo, day)
    except ValueError:
        return date(syr, smo, 1)


def sheet_month(title):
    t = norm(title).upper().replace("'", ' ').replace('.', ' ')
    for i, mo in enumerate(MONTHS):
        if t.startswith(mo[:3].upper()):
            m = re.search(r'(\d{2,4})', t)
            y = int(m.group(1)) if m else 0
            if y and y < 100:
                y += 2000
            return i + 1, (y or None)
    return None, None


def reading(v, key):
    """Numeric cell, dropped when it is outside a plausible range for that test."""
    n = num(v)
    if n is None:
        return None
    lo, hi = PLAUSIBLE[key]
    if not (lo <= n <= hi):
        EXCLUDED['readings'] += 1
        return None
    return n


def parse_quality_params(wb):
    recs = []
    last_year = 2025
    for ws in wb.worksheets:
        rows = sheet_rows(ws)
        if not rows:
            continue
        head = [norm(c).upper() for c in rows[0]]
        if 'AG CODES' in head or 'JOB NO' in head:
            idx = {}
            for c, h in enumerate(head):
                key = QP_MAP.get(h)
                if key and key not in idx:
                    idx[key] = c
            body = rows[1:]
        else:  # sheet saved without its header row
            idx = {k: c for c, k in enumerate(QP_DEFAULT_ORDER) if k}
            body = rows
        smo, syr = sheet_month(ws.title)
        syr = syr or last_year          # sheets like 'JULY' carry no year
        last_year = syr
        for row in body:
            def g(k):
                c = idx.get(k)
                return row[c] if c is not None and c < len(row) else None
            ag = norm(g('ag'))
            if not ag and not norm(g('job')):
                continue
            d = ag_date(ag, smo, syr)
            if d is None and smo:
                d = date(syr, smo, 1)
            if d is None:
                continue
            rec = {
                'd': iso(d),
                'sheet': ws.title,
                'job': norm(g('job')),
                'ag': ag,
                'product': norm(g('product')).upper(),
                'ph': reading(g('ph'), 'ph'),
                'salt': reading(g('salt'), 'salt'),
                'brix': reading(g('brix'), 'brix'),
                'vacuum': reading(g('vacuum'), 'vacuum'),
                'drain': reading(g('drain'), 'drain'),
                'nett': reading(g('nett'), 'nett'),
                'issues': norm(g('issues')).upper(),
            }
            if any(rec[k] is not None for k in ('ph', 'salt', 'brix', 'vacuum', 'drain')):
                recs.append(rec)
    recs.sort(key=lambda r: r['d'])
    return recs


def qp_trend(recs, tid, key, title, unit, note=''):
    """Daily mean of a canning parameter, with that day's spread."""
    byday = {}
    for r in recs:
        if r[key] is None:
            continue
        byday.setdefault(r['d'], []).append(r[key])
    pts = []
    for d in sorted(byday):
        vals = byday[d]
        pts.append({'d': d, 'v': {
            'avg': round(sum(vals) / len(vals), 3),
            'min': round(min(vals), 3),
            'max': round(max(vals), 3),
            'n': len(vals),
        }})
    series = [
        {'key': 'avg', 'label': 'Daily average', 'color': '#7c3aed'},
        {'key': 'min', 'label': 'Lowest', 'color': '#94a3b8', 'dashed': True},
        {'key': 'max', 'label': 'Highest', 'color': '#94a3b8', 'dashed': True},
    ]
    return {
        'id': tid, 'title': title, 'category': 'Canning parameters',
        'chart': 'line', 'unit': unit, 'series': series,
        'points': pts, 'stats': stats(pts, ['avg', 'min', 'max']), 'note': note,
    }


# ------------------------------------------------------------- categorical

def outcome(text):
    """Pass / fail / anything else (blank or 'NONE' means it was never recorded)."""
    t = norm(text).lower()
    if t.startswith('pass'):
        return 'passed'
    if t.startswith('fail'):
        return 'failed'
    return 'other'


def parse_allergen(wb):
    labels = {'Area': 'area', 'Result': 'result'}
    out = []
    for ws in wb.worksheets:
        rows = sheet_rows(ws)
        header_i = next((i for i, r in enumerate(rows[:6])
                         if any(norm(c).lower() == 'month' for c in r)), None)
        if header_i is None:
            continue
        header = rows[header_i]
        for c0 in [c for c, v in enumerate(header) if norm(v).lower() == 'month']:
            cols = {}
            for c in range(c0 + 2, min(c0 + 6, len(header))):
                lab = norm(header[c])
                if lab in labels:
                    cols[labels[lab]] = c
                elif not lab:
                    break
            if 'result' not in cols:
                continue
            for row in rows[header_i + 1:]:
                d = parse_date(row[c0 + 1]) if c0 + 1 < len(row) else None
                res = norm(row[cols['result']]) if cols['result'] < len(row) else ''
                if not d or not res:
                    continue
                area = norm(row[cols['area']]) if 'area' in cols and cols['area'] < len(row) else ''
                out.append({'d': iso(d), 'area': area.upper(), 'result': res.title()})
    out.sort(key=lambda r: r['d'])
    return out


def parse_emp(wb):
    log = []
    if 'Sampling Log' in wb.sheetnames:
        rows = sheet_rows(wb['Sampling Log'])
        hi = next((i for i, r in enumerate(rows[:15])
                   if r and norm(r[0]).lower().startswith('test date')), None)
        if hi is not None:
            head = [norm(c).lower() for c in rows[hi]]

            def col(*names):
                for n in names:
                    for c, h in enumerate(head):
                        if h.startswith(n):
                            return c
                return None
            ci = {k: col(*n) for k, n in {
                'd': ('test date',), 'zone': ('zone',), 'point': ('sample point',),
                'area': ('area',), 'test': ('test requested',), 'result': ('result',),
                'pf': ('pass/fail',), 'action': ('corrective',),
            }.items()}
            for row in rows[hi + 1:]:
                d = parse_date(row[ci['d']]) if ci['d'] is not None and ci['d'] < len(row) else None
                if not d:
                    continue

                def g(k):
                    c = ci.get(k)
                    return norm(row[c]) if c is not None and c < len(row) else ''
                log.append({'d': iso(d), 'zone': g('zone'), 'point': g('point'),
                            'area': g('area'), 'test': g('test'), 'result': g('result'),
                            'pf': g('pf').title(), 'action': g('action')})
    log.sort(key=lambda r: r['d'])

    trend = []
    if 'Trend Analysis' in wb.sheetnames:
        rows = sheet_rows(wb['Trend Analysis'])
        hi = next((i for i, r in enumerate(rows[:6])
                   if r and norm(r[0]).lower() == 'month'), None)
        if hi is not None:
            for row in rows[hi + 1:]:
                mo = norm(row[0])
                zone = norm(row[1]) if len(row) > 1 else ''
                if not mo or not zone:
                    continue
                trend.append({'month': mo, 'zone': zone, 'total': num(row[2]),
                              'pass': num(row[3]), 'fail': num(row[4]),
                              'failPct': num(row[5]) if len(row) > 5 else None})
    return log, trend


def parse_env(wb):
    years = {}
    for ws in wb.worksheets:
        if not re.match(r'^20\d\d$', norm(ws.title)):
            continue
        rows = sheet_rows(ws)
        if not rows:
            continue
        head = [norm(c) for c in rows[0]]
        mcols = [(c, norm(h)) for c, h in enumerate(head) if norm(h) in MONTHS]
        items = []
        for row in rows[1:]:
            name = norm(row[0])
            if not name:
                continue
            vals = {}
            for c, mo in mcols:
                n = num(row[c]) if c < len(row) else None
                if n is not None:
                    vals[mo] = n
            if vals:
                items.append({'test': name, 'months': vals})
        if items:
            years[norm(ws.title)] = items
    return years


# -------------------------------------------------------------------- main

def main():
    print('Reading workbooks from: %s' % SRC)
    data = {
        'generatedAt': datetime.now().isoformat(timespec='seconds'),
        'sources': [],
        'trends': [],
        'allergen': [],
        'qualityParams': [],
        'emp': {'log': [], 'trend': []},
        'env': {},
    }

    def src(name, stamp):
        if stamp:
            data['sources'].append({'file': name, 'modified': stamp})

    # --- temperature / pH / humidity monitoring ---------------------------
    wb, stamp = load(F_TEMP)
    if wb:
        src(F_TEMP, stamp)
        chill_labels = {'Chiller 1 Temp': 'c1', 'Chiller 2 Temp': 'c2'}
        chill_series = [{'key': 'c1', 'label': 'Chiller 1', 'color': '#2563eb'},
                        {'key': 'c2', 'label': 'Chiller 2', 'color': '#0891b2'}]
        for name, tid, title in [
            ('Chiller Monitoring 2025', 'chiller-2025', 'Chiller temperature 2025'),
            ('Chiller Monitoring 2026', 'chiller-2026', 'Chiller temperature 2026'),
        ]:
            if name in wb.sheetnames:
                t = monitoring_trend(wb[name], tid, title, '°C',
                                     chill_labels, chill_series)
                t['spec'] = {'max': 10, 'label': 'Max 10 °C'}
                data['trends'].append(t)

        ph_labels = {'pH: Fresh water': 'fresh', 'pH: Salt water': 'salt', 'Salt %': 'saltPct'}
        for name, tid, title in [
            ('Water pH Monitoring 2025', 'water-ph-2025', 'Water pH 2025'),
            ('Water pH Monitoring 2026', 'water-ph-2026', 'Water pH 2026'),
        ]:
            if name in wb.sheetnames:
                data['trends'].append(monitoring_trend(
                    wb[name], tid, title, 'pH', ph_labels,
                    [{'key': 'fresh', 'label': 'Fresh water pH', 'color': '#2563eb'},
                     {'key': 'salt', 'label': 'Salt water pH', 'color': '#0d9488'}]))
                data['trends'].append(monitoring_trend(
                    wb[name], tid.replace('-ph-', '-salinity-'),
                    title.replace('pH', 'salinity'), '%', ph_labels,
                    [{'key': 'saltPct', 'label': 'Salt %', 'color': '#ea580c'}]))

        if 'Wet Storage (LHA)' in wb.sheetnames:
            t = monitoring_trend(wb['Wet Storage (LHA)'], 'lha-temp',
                                 'Wet storage (LHA) water temperature', '°C',
                                 {'Temp': 'temp'},
                                 [{'key': 'temp', 'label': 'LHA water temp', 'color': '#0284c7'}])
            t['spec'] = {'min': 12, 'max': 18, 'label': 'Spec 12–18 °C'}
            data['trends'].append(t)

        if 'Dry Room Monitoring' in wb.sheetnames:
            t = monitoring_trend(wb['Dry Room Monitoring'], 'dry-room',
                                 'Dry room temperature & humidity', '°C / %RH',
                                 {'Temperature': 'temp', 'Humidity': 'rh'},
                                 [{'key': 'temp', 'label': 'Temperature', 'color': '#b45309'},
                                  {'key': 'rh', 'label': 'Humidity', 'color': '#0891b2'}])
            t['note'] = ('Plotted exactly as captured in the workbook. Several readings sit '
                         'outside a plausible range for a room log (humidity above 100 %), so '
                         'confirm against the source records before quoting these figures.')
            data['trends'].append(t)
        wb.close()

    # --- canning quality parameters ---------------------------------------
    wb, stamp = load(F_QP)
    if wb:
        src(F_QP, stamp)
        recs = parse_quality_params(wb)
        data['qualityParams'] = recs
        data['trends'] += [
            qp_trend(recs, 'qp-ph', 'ph', 'Canned product pH', 'pH'),
            qp_trend(recs, 'qp-salt', 'salt', 'Canned product salt', '%'),
            qp_trend(recs, 'qp-brix', 'brix', 'Canned product brix', '°Bx'),
            qp_trend(recs, 'qp-vacuum', 'vacuum', 'Can vacuum', 'kPa'),
            qp_trend(recs, 'qp-drain', 'drain', 'Drained mass', 'g'),
        ]
        bym = {}
        for r in recs:
            b = bym.setdefault(r['d'][:7], {'checks': 0, 'issues': 0})
            b['checks'] += 1
            if r['issues'] and r['issues'] not in ('NONE', 'N/A', '-', '0'):
                b['issues'] += 1
        pts = [{'d': mo + '-01', 'v': {
            'checks': bym[mo]['checks'], 'issues': bym[mo]['issues'],
            'pct': round(100.0 * bym[mo]['issues'] / bym[mo]['checks'], 2)}}
            for mo in sorted(bym)]
        data['trends'].append({
            'id': 'qp-issues', 'title': 'QC issues raised per month',
            'category': 'Canning parameters', 'chart': 'bar', 'unit': 'checks',
            'series': [{'key': 'checks', 'label': 'Cans checked', 'color': '#cbd5e1'},
                       {'key': 'issues', 'label': 'With a QC issue', 'color': '#dc2626'}],
            'points': pts, 'stats': stats(pts, ['checks', 'issues', 'pct']),
            'note': 'A check counts as an issue when the QC Issues column reads anything other than NONE.',
        })
        wb.close()

    # --- allergen swabs -----------------------------------------------------
    wb, stamp = load(F_ALLERGEN)
    if wb:
        src(F_ALLERGEN, stamp)
        data['allergen'] = parse_allergen(wb)
        bym = {}
        for r in data['allergen']:
            b = bym.setdefault(r['d'][:7], {'passed': 0, 'failed': 0, 'other': 0})
            b[outcome(r['result'])] += 1
        pts = [{'d': mo + '-01', 'v': bym[mo]} for mo in sorted(bym)]
        data['trends'].append({
            'id': 'allergen', 'title': 'Allergen swab results',
            'category': 'Verification', 'chart': 'bar', 'unit': 'swabs',
            'series': [{'key': 'passed', 'label': 'Passed', 'color': '#16a34a'},
                       {'key': 'failed', 'label': 'Failed', 'color': '#dc2626'},
                       {'key': 'other', 'label': 'Not recorded', 'color': '#cbd5e1'}],
            'points': pts, 'stats': stats(pts, ['passed', 'failed', 'other']),
            'note': 'Swab results as recorded in the allergen monitoring workbook.',
        })
        wb.close()

    # --- environmental monitoring programme ---------------------------------
    wb, stamp = load(F_EMP)
    if wb:
        src(F_EMP, stamp)
        log, trend = parse_emp(wb)
        data['emp'] = {'log': log, 'trend': trend}
        bym = {}
        for r in log:
            b = bym.setdefault(r['d'][:7], {'passed': 0, 'failed': 0, 'other': 0})
            b[outcome(r['pf'])] += 1
        pts = [{'d': mo + '-01', 'v': bym[mo]} for mo in sorted(bym)]
        data['trends'].append({
            'id': 'emp', 'title': 'EMP swab results (REC 2.5.7.1)',
            'category': 'Verification', 'chart': 'bar', 'unit': 'samples',
            'series': [{'key': 'passed', 'label': 'Pass', 'color': '#16a34a'},
                       {'key': 'failed', 'label': 'Fail', 'color': '#dc2626'},
                       {'key': 'other', 'label': 'Not recorded', 'color': '#cbd5e1'}],
            'points': pts, 'stats': stats(pts, ['passed', 'failed', 'other']),
            'note': 'From the EMP Sampling Log. The zone breakdown is in the table below.',
        })
        wb.close()

    wb, stamp = load(F_ENV)
    if wb:
        src(F_ENV, stamp)
        data['env'] = parse_env(wb)
        wb.close()

    data['trends'] = [t for t in data['trends'] if t['points']]
    data['excluded'] = dict(EXCLUDED)

    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(data, fh, separators=(',', ':'))

    print('Wrote %s (%.0f KB)' % (OUT, os.path.getsize(OUT) / 1024.0))
    for t in data['trends']:
        print('  %-20s %-44s %4d points' % (t['id'], t['title'], len(t['points'])))
    print('  allergen swabs: %d | EMP samples: %d | canning checks: %d'
          % (len(data['allergen']), len(data['emp']['log']), len(data['qualityParams'])))
    print('  excluded as implausible: %d dates, %d readings'
          % (EXCLUDED['dates'], EXCLUDED['readings']))


if __name__ == '__main__':
    main()
